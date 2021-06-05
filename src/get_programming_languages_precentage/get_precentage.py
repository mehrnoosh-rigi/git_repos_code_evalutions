import csv
import codecs
import os
import pathlib
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

import xlsxwriter


def evaluate_cloc_file(file):
    print(file)


def run(cloc_directory, tool):
    # workbook = xlsxwriter.Workbook(f"results/precentage_in_excel/{tool}.xlsx")
    # worksheet = workbook.add_worksheet()
    # worksheet.set_column('A:A', 40)
    i = 1
    result_file_dir = f"results/precentage_in_excel/{tool}.xlsx"
    for path in pathlib.Path(cloc_directory).iterdir():
        with open(path) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            print("path", path)
            # lines_numbers = len(list(csv_reader)) - 1
            # with open(f"results/precentage_in_excel/{tool}.xlsx", "a") as result_file:
            # result_file.write("--------------------------")
            # result_file.write(path.name)
            # result_file.write("\n")
            result_of_language = []
            result_of_counter = []
            for row in csv_reader:
                # csv_file = list(csv.reader)
                # print(csv_file)
                # line = csv_file[i+1]
                language = row[1]
                count = row[4]
                result_of_language.append(row[1])
                result_of_counter.append(row[4])
                print(language, count)
                # with open(f"results/precentage_in_excel/{tool}.txt", "a") as result_file:
                #     result_file.write(language)
                # result_file.write("--->")
                # result_file.write(count)
                # result_file.write("\n")

                # results.append({language, count, "\n"})
                # worksheet.write(i, 1, language, count)

            # "Name Of Repo": path.name,

            df = pd.DataFrame({"Language": result_of_language,
                               "Lines of code counter": result_of_counter})

            # Create a Pandas Excel writer using XlsxWriter as the engine.
            writer = pd.ExcelWriter(f"results/precentage_in_excel/{tool}/{path.name}.xlsx", engine='xlsxwriter')

            # Convert the dataframe to an XlsxWriter Excel object.
            df.to_excel(writer, sheet_name=tool, index=False)
            workbook = writer.book
            worksheet = writer.sheets[tool]
            chart = workbook.add_chart({'type': 'column'})
            chart.add_series({"values": f"={tool}!$B$2:$B$5"})
            worksheet.insert_chart('D2', chart)

            # Close the Pandas Excel writer and output the Excel file.
            writer.save()
            # df = pd.DataFrame({"Language": result_of_language,
            #                    "Lines of code counter": result_of_counter})
            # writer = pd.ExcelWriter(result_file_dir, engine='openpyxl')
            # # try to open an existing workbook
            # writer.book = load_workbook("results/precentage_in_excel/selenium.xlsx")
            # # copy existing sheets
            # writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            # # read existing file
            # reader = pd.read_excel(rf"results/precentage_in_excel/{tool}.xlsx")
            # # write out the new sheet
            # df.to_excel(writer, index=False, header=False, startrow=len(reader) + 1)
            #
            # writer.close()

            i += 1
            # total_lines_of_code = csv_reader[lines_numbers+1]
            # worksheet.write(lines_numbers, total_lines_of_code)
    # with open(f"results/precentage_in_excel/{tool}.txt", "a") as result_file:
    #     result_file.write(results)

    # workbook.close()

    # print("path", path)
    # print("current", os.getcwd())
    #
    # entries = Path(path)
    # for entry in entries.iterdir():
    #     try:
    #         cloc_file = f"{path}/cloc/{entry.name}"
    #         with open(cloc_file) as f:
    #             print("here")
    #     except:
    #         print("not cloc file")
